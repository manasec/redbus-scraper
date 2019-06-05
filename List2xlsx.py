import openpyxl
from html_table_extractor.extractor import Extractor
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import re

class List2xlsx(object):
    
    @classmethod
    def list2xlsx(self, workbook, sheetname, alist):
        wb = openpyxl.load_workbook(workbook)
        sheet = wb[sheetname]
        if sheet.max_row>1:
            next_row = sheet.max_row+2
        else:
            next_row = 1
            alist.insert(0,['platform', 'coupon_code', 'offer_type', 'details', 'mini',
                   'channel', 'applicable_bw','validity[index]','constraint', 'link'])
        for i in range(len(alist)):
            for j in range(len(alist[0])):
                sheet.cell(row=next_row+i, column=j+1).value = alist[i][j]
        
        wb.save(workbook)
    

    @classmethod
    def formatting(self, workbook, sheetname):
        wb = openpyxl.load_workbook(workbook)
        sheet = wb[sheetname]
        fontObj1 = Font(bold=True)
        for col in range(1,sheet.max_column+1):
            sheet.cell(row=1, column=col).font = fontObj1
        for col in sheet.columns:
             max_length = 0
             column = col[0].column
             
             for cell in col:
                 try: 
                     if len(str(cell.value)) > max_length:
                         max_length = len(cell.value)
                 except:
                     pass
             adjusted_width = max_length + 2
             sheet.column_dimensions[get_column_letter(column)].width = adjusted_width       
        wb.save(workbook)
